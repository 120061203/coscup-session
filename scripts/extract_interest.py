#!/usr/bin/env python3
"""Extract session id -> 感興趣程度 from the COSCUP excel into src/data/interestSeed.json.

Usage: python3 scripts/extract_interest.py <path-to-xlsx>
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else next(ROOT.glob("*.xlsx"))
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    seed = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        link, interest = row[8], row[9]
        if not link:
            continue
        match = re.search(r"/session/([A-Za-z0-9]+)", str(link))
        if not match:
            continue
        seed[match.group(1)] = int(interest) if interest is not None else 0

    out_path = ROOT / "src" / "data" / "interestSeed.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(seed)} entries to {out_path}")


if __name__ == "__main__":
    main()
